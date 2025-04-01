"""
Feasibility Reporter Module for Waffle Production Optimization.

This module generates concise feasibility reports and identifies critical
data points in input files for review.
"""
from typing import Dict, List, Any, Optional, Tuple
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import zipfile
import io
from feasibility_check import FeasibilityChecker

class FeasibilityReporter:
    def __init__(self, checker: FeasibilityChecker, input_files: Dict[str, str] = None):
        """
        Initialize the feasibility reporter.
        
        Args:
            checker: An initialized FeasibilityChecker instance
            input_files: Dictionary of input file names and paths
        """
        self.checker = checker
        self.input_files = input_files or {}
        self.critical_data_points = []
        
    def generate_concise_report(self) -> Dict[str, Any]:
        """
        Generate a concise text report of the feasibility assessment.
        
        Returns:
            Dict containing report sections and overall status
        """
        is_feasible = len(self.checker.feasibility_issues) == 0
        has_warnings = len(self.checker.warnings) > 0
        
        # Calculate key metrics
        total_demand = sum(self.checker.data['total_demand'].values())
        total_supply = sum(self.checker.data['total_supply'].values())
        buffer_percentage = (total_supply / total_demand * 100) if total_demand > 0 else float('inf')
        
        # Find problematic waffles
        waffle_types = self.checker.data['waffle_types']
        pan_types = self.checker.data['pan_types']
        allowed = self.checker.data['allowed']
        
        incompatible_waffles = []
        for waffle in waffle_types:
            compatible_pans = [p for p in pan_types if allowed.get((waffle, p), False)]
            if len(compatible_pans) <= 0:
                incompatible_waffles.append(waffle)
            elif len(compatible_pans) <= 2:  # Few compatible pans is also a potential issue
                incompatible_waffles.append(waffle)
        
        # Identify critical weeks with supply issues
        weeks = sorted(self.checker.data['weeks'])
        critical_weeks = []
        tight_weeks = []
        
        for week in weeks:
            week_demand = self.checker.data['total_demand'].get(week, 0)
            week_supply = self.checker.data['total_supply'].get(week, 0)
            
            if week_demand > 0:
                week_ratio = (week_supply / week_demand) * 100
                if week_supply < week_demand:
                    critical_weeks.append({
                        'week': week,
                        'supply': week_supply,
                        'demand': week_demand,
                        'ratio': week_ratio
                    })
                elif week_supply < week_demand * 1.1:  # Less than 10% buffer
                    tight_weeks.append({
                        'week': week,
                        'supply': week_supply,
                        'demand': week_demand,
                        'ratio': week_ratio
                    })
        
        # Identify compatibility-constrained supply issues
        compatibility_issues = []
        for issue in self.checker.feasibility_issues:
            if "Insufficient compatible pan supply for waffle" in issue:
                compatibility_issues.append(issue)
        
        # Count unique waffles with compatibility supply issues
        waffles_with_compat_issues = set()
        for issue in compatibility_issues:
            # Extract waffle name from the issue message
            try:
                waffle_name = issue.split("waffle '")[1].split("'")[0]
                waffles_with_compat_issues.add(waffle_name)
            except:
                pass
        
        # Group consecutive critical weeks for concise reporting
        grouped_critical_weeks = self._group_consecutive_weeks(critical_weeks)
        grouped_tight_weeks = self._group_consecutive_weeks(tight_weeks)
        
        # Create report data structure
        report = {
            'status': 'FEASIBLE' if is_feasible and not has_warnings else 
                     ('FEASIBLE WITH WARNINGS' if is_feasible else 'INFEASIBLE'),
            'metrics': {
                'total_supply': total_supply,
                'total_demand': total_demand,
                'buffer_percentage': buffer_percentage,
                'all_compatible': len(incompatible_waffles) == 0,
                'critical_week_count': len(critical_weeks),
                'tight_week_count': len(tight_weeks),
                'total_issues': len(self.checker.feasibility_issues),
                'total_warnings': len(self.checker.warnings),
                'compatibility_issues_count': len(compatibility_issues),
                'waffles_with_compat_issues': len(waffles_with_compat_issues)
            },
            'critical_weeks': grouped_critical_weeks,
            'tight_weeks': grouped_tight_weeks,
            'incompatible_waffles': incompatible_waffles,
            'compatibility_issues': compatibility_issues,
            'issues': self.checker.feasibility_issues,
            'warnings': self.checker.warnings,
            'recommendations': self._generate_recommendations(critical_weeks, tight_weeks, incompatible_waffles, compatibility_issues)
        }
        
        return report
    
    def generate_text_report(self) -> str:
        """
        Generate a concise text-based report for display in a single window.
        
        Returns:
            String with formatted text report
        """
        report_data = self.generate_concise_report()
        status = report_data['status']
        metrics = report_data['metrics']
        
        # Create the header
        if status == 'FEASIBLE':
            status_icon = '✓'
            status_display = f"{status_icon} FEASIBLE"
        elif status == 'FEASIBLE WITH WARNINGS':
            status_icon = '⚠️'
            status_display = f"{status_icon} FEASIBLE WITH WARNINGS"
        else:
            status_icon = '✗'
            status_display = f"{status_icon} INFEASIBLE"
        
        # Start building the report text
        report_text = "=== FEASIBILITY ASSESSMENT SUMMARY ===\n"
        report_text += f"STATUS: {status_display}\n\n"
        
        # Add critical metrics section
        report_text += "CRITICAL METRICS:\n"
        
        # Supply vs demand
        buffer_icon = '✓' if metrics['buffer_percentage'] >= 110 else ('⚠️' if metrics['buffer_percentage'] >= 100 else '✗')
        report_text += f"{buffer_icon} Total Supply: {metrics['total_supply']:,} pans ({metrics['buffer_percentage']:.1f}% of demand)\n"
        
        # Waffle compatibility
        compat_icon = '✓' if metrics['all_compatible'] else '✗'
        if metrics['all_compatible']:
            report_text += f"{compat_icon} All waffle types have compatible pans\n"
        else:
            report_text += f"{compat_icon} {len(report_data['incompatible_waffles'])} waffle types have limited or no compatible pans\n"
        
        # Weekly supply
        weekly_icon = '✓' if metrics['critical_week_count'] == 0 and metrics['tight_week_count'] == 0 else ('⚠️' if metrics['critical_week_count'] == 0 else '✗')
        issue_weeks = metrics['critical_week_count'] + metrics['tight_week_count']
        if issue_weeks > 0:
            report_text += f"{weekly_icon} Weekly Supply Issues: {issue_weeks} weeks with supply below 110% buffer\n"
        else:
            report_text += f"{weekly_icon} All weeks have adequate supply (>110% of demand)\n"
        
        # Compatibility-constrained supply issues
        compat_supply_icon = '✓' if metrics['compatibility_issues_count'] == 0 else '✗'
        if metrics['compatibility_issues_count'] > 0:
            report_text += f"{compat_supply_icon} Compatibility-Constrained Supply: {metrics['waffles_with_compat_issues']} waffle types have insufficient compatible pan supply\n"
        else:
            report_text += f"{compat_supply_icon} All waffle types have sufficient compatible pan supply\n"
        
        # Add critical weeks section if there are any
        if metrics['critical_week_count'] > 0:
            report_text += "\nCRITICAL WEEKS:\n"
            for group in report_data['critical_weeks']:
                if len(group) == 1:
                    week = group[0]
                    ratio_pct = week['ratio']
                    report_text += f"✗ Week {week['week']}: Supply {week['supply']:,} ({ratio_pct:.1f}% of demand)\n"
                else:
                    start_week = group[0]['week']
                    end_week = group[-1]['week']
                    avg_ratio = sum(w['ratio'] for w in group) / len(group)
                    report_text += f"✗ Week {start_week}-{end_week}: Supply < {avg_ratio:.1f}% of demand\n"
        
        # Add compatibility-constrained supply issues if there are any
        if metrics['compatibility_issues_count'] > 0:
            report_text += "\nCOMPATIBILITY-CONSTRAINED SUPPLY ISSUES:\n"
            # Limit to a few examples for conciseness
            max_issues = min(5, metrics['compatibility_issues_count'])
            for i, issue in enumerate(report_data['compatibility_issues'][:max_issues]):
                # Extract and format the key information
                waffle_name = issue.split("waffle '")[1].split("'")[0]
                
                if "weeks" in issue:
                    # Range of weeks issue
                    weeks_range = issue.split("in weeks ")[1].split(":")[0]
                    avg_deficit = issue.split("Average ")[1].split("%")[0]
                    report_text += f"✗ Waffle {waffle_name}: Weeks {weeks_range} - {avg_deficit}% shortage of compatible pans\n"
                else:
                    # Single week issue
                    week = issue.split("in week ")[1].split(":")[0]
                    deficit_pct = issue.split("(")[1].split("%")[0]
                    report_text += f"✗ Waffle {waffle_name}: Week {week} - {deficit_pct}% shortage of compatible pans\n"
                
            # Add indication if there are more issues
            if metrics['compatibility_issues_count'] > max_issues:
                remaining = metrics['compatibility_issues_count'] - max_issues
                report_text += f"  ... and {remaining} more compatibility issues\n"
        
        # Add recommendations
        if report_data['recommendations']:
            report_text += "\nRECOMMENDATIONS:\n"
            for rec in report_data['recommendations']:
                report_text += f"• {rec}\n"
        
        # Add data issues section
        data_points = len(self.critical_data_points)
        if data_points > 0:
            report_text += f"\nDATA ISSUES: {data_points} potential problems in input files\n"
            report_text += "[DOWNLOAD MARKED FILES]\n"
        
        return report_text
    
    def identify_critical_data(self) -> List[Dict[str, Any]]:
        """
        Identify critical data points in input files that need review.
        
        Returns:
            List of dictionaries with critical data point information
        """
        critical_points = []
        report_data = self.generate_concise_report()
        
        # Process critical weeks - mark them in supply file
        for group in report_data['critical_weeks']:
            for week_data in group:
                week = week_data['week']
                critical_points.append({
                    'file': 'PanSupply.xlsx',
                    'sheet': 'Weekly Supply',  # Assumed sheet name
                    'cell_type': 'row',
                    'identifier': week,
                    'message': f"Week {week} supply ({week_data['supply']:,}) is only {week_data['ratio']:.1f}% of demand ({week_data['demand']:,})"
                })
        
        # Process tight weeks with warning
        for group in report_data['tight_weeks']:
            for week_data in group:
                week = week_data['week']
                critical_points.append({
                    'file': 'PanSupply.xlsx',
                    'sheet': 'Weekly Supply',
                    'cell_type': 'row',
                    'identifier': week,
                    'message': f"Week {week} supply is tight: {week_data['ratio']:.1f}% of demand",
                    'is_warning': True  # This is a warning, not critical
                })
        
        # Process incompatible waffles - mark them in combinations file
        for waffle in report_data['incompatible_waffles']:
            critical_points.append({
                'file': 'WafflePanCombinations.xlsx',
                'sheet': 'Combinations',  # Assumed sheet name
                'cell_type': 'waffle',
                'identifier': waffle,
                'message': f"Waffle {waffle} has limited or no compatible pans"
            })
        
        self.critical_data_points = critical_points
        return critical_points
    
    def generate_marked_excel_files(self, output_dir: str = '.') -> Dict[str, str]:
        """
        Generate marked versions of Excel files highlighting critical data.
        
        Args:
            output_dir: Directory to save marked files
            
        Returns:
            Dictionary mapping original filenames to marked filenames
        """
        if not self.input_files:
            return {}
            
        if not self.critical_data_points:
            self.identify_critical_data()
            
        marked_files = {}
        
        for orig_filename, filepath in self.input_files.items():
            if not os.path.exists(filepath):
                continue
                
            # Find critical points for this file
            file_points = [p for p in self.critical_data_points if p['file'] == orig_filename]
            
            if not file_points:
                continue
                
            # Load the workbook
            try:
                wb = load_workbook(filepath)
                
                # Track if any changes were made
                changes_made = False
                
                # Process each sheet
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    # Find points for this sheet
                    sheet_points = [p for p in file_points if p['sheet'] == sheet_name]
                    
                    if not sheet_points:
                        continue
                        
                    # Process each point
                    for point in sheet_points:
                        # Find the cells to mark based on point type
                        cells_to_mark = self._locate_cells(sheet, point)
                        is_warning = point.get('is_warning', False)
                        
                        # Apply formatting to cells
                        if cells_to_mark:
                            changes_made = True
                            for cell in cells_to_mark:
                                # Choose colors based on warning vs critical
                                if is_warning:
                                    fill_color = "FFFFCC"  # Light yellow for warnings
                                    border_color = "FFD700"  # Gold
                                else:
                                    fill_color = "FFCCCC"  # Light red for critical issues
                                    border_color = "FF0000"  # Red
                                
                                # Apply formatting
                                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                                
                                # Add borders
                                thin_border = Border(
                                    left=Side(style='thin', color=border_color),
                                    right=Side(style='thin', color=border_color),
                                    top=Side(style='thin', color=border_color),
                                    bottom=Side(style='thin', color=border_color)
                                )
                                cell.border = thin_border
                            
                            # Add comment to the first cell
                            if cells_to_mark:
                                first_cell = cells_to_mark[0]
                                comment = Comment(point['message'], "Feasibility Check")
                                comment.width = 300
                                comment.height = 100
                                first_cell.comment = comment
                
                if changes_made:
                    # Save as a new file
                    marked_filename = orig_filename.replace('.xlsx', '_marked.xlsx')
                    marked_filepath = os.path.join(output_dir, marked_filename)
                    wb.save(marked_filepath)
                    marked_files[orig_filename] = marked_filepath
            
            except Exception as e:
                print(f"Error marking file {orig_filename}: {str(e)}")
                
        return marked_files
    
    def create_zip_with_marked_files(self, output_path: str = 'marked_files.zip') -> str:
        """
        Create a ZIP file containing all marked Excel files.
        
        Args:
            output_path: Path to save the ZIP file
            
        Returns:
            Path to the created ZIP file
        """
        marked_files = self.generate_marked_excel_files()
        
        if not marked_files:
            return ""
            
        try:
            with zipfile.ZipFile(output_path, 'w') as zipf:
                for orig_file, marked_path in marked_files.items():
                    zipf.write(marked_path, os.path.basename(marked_path))
                    
                # Also add a text report
                report_text = self.generate_text_report()
                zipf.writestr('feasibility_report.txt', report_text)
                
            return output_path
        
        except Exception as e:
            print(f"Error creating ZIP file: {str(e)}")
            return ""
    
    def generate_html_report(self, output_path: str = 'feasibility_report.html', 
                            include_download_link: bool = True) -> str:
        """
        Generate an HTML report with the feasibility assessment.
        
        Args:
            output_path: Path to save the HTML report
            include_download_link: Whether to include a download link for marked files
            
        Returns:
            Path to the generated HTML file
        """
        report_data = self.generate_concise_report()
        
        # Create the HTML report
        html = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Waffle Optimizer - Feasibility Report</title>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    margin: 20px;
                    max-width: 1000px;
                    margin: 0 auto;
                }
                h1 {
                    color: #333;
                    border-bottom: 2px solid #ddd;
                    padding-bottom: 10px;
                }
                .status-container {
                    padding: 15px;
                    border-radius: 5px;
                    margin-bottom: 20px;
                    font-weight: bold;
                    font-size: 1.2em;
                    text-align: center;
                }
                .feasible {
                    background-color: #dff0d8;
                    color: #3c763d;
                    border: 1px solid #d6e9c6;
                }
                .warning {
                    background-color: #fcf8e3;
                    color: #8a6d3b;
                    border: 1px solid #faebcc;
                }
                .infeasible {
                    background-color: #f2dede;
                    color: #a94442;
                    border: 1px solid #ebccd1;
                }
                .metrics-container {
                    display: grid;
                    grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
                    gap: 15px;
                    margin-bottom: 20px;
                }
                .metric-box {
                    background-color: #f9f9f9;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    padding: 15px;
                }
                .metric-box h3 {
                    margin-top: 0;
                    border-bottom: 1px solid #eee;
                    padding-bottom: 5px;
                }
                .section {
                    margin-bottom: 30px;
                }
                .good {
                    color: #3c763d;
                }
                .warning-text {
                    color: #8a6d3b;
                }
                .critical {
                    color: #a94442;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin-bottom: 20px;
                }
                table, th, td {
                    border: 1px solid #ddd;
                }
                th, td {
                    padding: 10px;
                    text-align: left;
                }
                th {
                    background-color: #f2f2f2;
                }
                tr:nth-child(even) {
                    background-color: #f9f9f9;
                }
                .recommendation {
                    background-color: #e8f4f8;
                    border-left: 5px solid #5bc0de;
                    padding: 10px 15px;
                    margin-bottom: 10px;
                }
                .download-button {
                    display: inline-block;
                    padding: 10px 15px;
                    background-color: #337ab7;
                    color: white;
                    text-decoration: none;
                    border-radius: 5px;
                    font-weight: bold;
                    margin-top: 20px;
                }
                .download-button:hover {
                    background-color: #286090;
                }
                .icon {
                    font-weight: bold;
                    margin-right: 5px;
                }
                pre {
                    background-color: #f5f5f5;
                    padding: 15px;
                    border-radius: 5px;
                    overflow-x: auto;
                    font-family: monospace;
                    white-space: pre-wrap;
                }
            </style>
        </head>
        <body>
            <h1>Waffle Production Feasibility Assessment</h1>
        """
        
        # Add status section
        status_class = "feasible" if report_data['status'] == 'FEASIBLE' else (
                       "warning" if report_data['status'] == 'FEASIBLE WITH WARNINGS' else "infeasible")
        status_icon = "✓" if report_data['status'] == 'FEASIBLE' else (
                      "⚠️" if report_data['status'] == 'FEASIBLE WITH WARNINGS' else "✗")
        
        html += f"""
            <div class="status-container {status_class}">
                <span class="icon">{status_icon}</span> {report_data['status']}
            </div>
        """
        
        # Add metrics section
        metrics = report_data['metrics']
        buffer_class = "good" if metrics['buffer_percentage'] >= 110 else (
                      "warning-text" if metrics['buffer_percentage'] >= 100 else "critical")
        
        html += """
            <div class="section">
                <h2>Critical Metrics</h2>
                <div class="metrics-container">
        """
        
        # Supply vs Demand metric
        html += f"""
            <div class="metric-box">
                <h3>Supply vs Demand</h3>
                <p>Total Supply: <span class="{buffer_class}">{metrics['total_supply']:,}</span> pans</p>
                <p>Total Demand: {metrics['total_demand']:,} pans</p>
                <p>Buffer: <span class="{buffer_class}">{metrics['buffer_percentage']:.1f}%</span></p>
            </div>
        """
        
        # Waffle Compatibility metric
        compat_class = "good" if metrics['all_compatible'] else "critical"
        incompatible_count = len(report_data['incompatible_waffles'])
        
        html += f"""
            <div class="metric-box">
                <h3>Waffle Compatibility</h3>
                <p>Status: <span class="{compat_class}">
                    {incompatible_count} waffle types with limited/no compatible pans
                </span></p>
            </div>
        """
        
        # Weekly Supply Issues metric
        weekly_class = "good" if metrics['critical_week_count'] == 0 else "critical"
        
        html += f"""
            <div class="metric-box">
                <h3>Weekly Supply Issues</h3>
                <p>Critical Weeks: <span class="{weekly_class}">{metrics['critical_week_count']}</span></p>
                <p>Weeks with Tight Supply: <span class="warning-text">{metrics['tight_week_count']}</span></p>
            </div>
        """
        
        # Compatibility-Constrained Supply metric
        compat_supply_class = "good" if metrics['compatibility_issues_count'] == 0 else "critical"
        
        html += f"""
            <div class="metric-box">
                <h3>Compatible Supply Issues</h3>
                <p>Waffles with Issues: <span class="{compat_supply_class}">{metrics['waffles_with_compat_issues']}</span></p>
                <p>Total Issues: <span class="{compat_supply_class}">{metrics['compatibility_issues_count']}</span></p>
            </div>
        """
        
        html += """
                </div>
            </div>
        """
        
        # Add critical weeks section if any
        if report_data['critical_weeks']:
            html += """
            <div class="section">
                <h2>Critical Weeks</h2>
                <table>
                    <tr>
                        <th>Week(s)</th>
                        <th>Supply</th>
                        <th>Demand</th>
                        <th>Supply/Demand Ratio</th>
                    </tr>
            """
            
            for group in report_data['critical_weeks']:
                if len(group) == 1:
                    week_data = group[0]
                    html += f"""
                    <tr>
                        <td>Week {week_data['week']}</td>
                        <td>{week_data['supply']:,}</td>
                        <td>{week_data['demand']:,}</td>
                        <td class="critical">{week_data['ratio']:.1f}%</td>
                    </tr>
                    """
                else:
                    # Grouped weeks
                    start_week = group[0]['week']
                    end_week = group[-1]['week']
                    total_supply = sum(w['supply'] for w in group)
                    total_demand = sum(w['demand'] for w in group)
                    avg_ratio = sum(w['ratio'] for w in group) / len(group)
                    
                    html += f"""
                    <tr>
                        <td>Weeks {start_week}-{end_week}</td>
                        <td>{total_supply:,}</td>
                        <td>{total_demand:,}</td>
                        <td class="critical">{avg_ratio:.1f}%</td>
                    </tr>
                    """
            
            html += """
                </table>
            </div>
            """
        
        # Add incompatible waffles section if any
        if report_data['incompatible_waffles']:
            html += """
            <div class="section">
                <h2>Waffles with Limited Compatibility</h2>
                <ul>
            """
            
            for waffle in report_data['incompatible_waffles']:
                html += f"""
                <li class="critical">Waffle {waffle}</li>
                """
            
            html += """
                </ul>
            </div>
            """
        
        # Add compatibility-constrained supply issues section if any
        if report_data['compatibility_issues']:
            html += """
            <div class="section">
                <h2>Compatibility-Constrained Supply Issues</h2>
                <table>
                    <tr>
                        <th>Waffle Type</th>
                        <th>Issue</th>
                    </tr>
            """
            
            # Group issues by waffle type for cleaner display
            waffle_issues = {}
            for issue in report_data['compatibility_issues']:
                try:
                    waffle_name = issue.split("waffle '")[1].split("'")[0]
                    if waffle_name not in waffle_issues:
                        waffle_issues[waffle_name] = []
                    waffle_issues[waffle_name].append(issue)
                except:
                    pass
            
            for waffle, issues in waffle_issues.items():
                html += f"""
                <tr>
                    <td rowspan="{len(issues)}">Waffle {waffle}</td>
                """
                
                # Add first issue
                issue_text = issues[0].split(f"waffle '{waffle}'")[1].strip()
                html += f"""
                    <td class="critical">{issue_text}</td>
                </tr>
                """
                
                # Add remaining issues
                for issue in issues[1:]:
                    issue_text = issue.split(f"waffle '{waffle}'")[1].strip()
                    html += f"""
                <tr>
                    <td class="critical">{issue_text}</td>
                </tr>
                """
            
            html += """
                </table>
            </div>
            """
        
        # Add recommendations section
        if report_data['recommendations']:
            html += """
            <div class="section">
                <h2>Recommendations</h2>
            """
            
            for rec in report_data['recommendations']:
                html += f"""
                <div class="recommendation">{rec}</div>
                """
            
            html += """
            </div>
            """
        
        # Add raw text report section
        text_report = self.generate_text_report()
        html += """
            <div class="section">
                <h2>Concise Text Report</h2>
                <pre>
        """
        html += text_report
        html += """
                </pre>
            </div>
        """
        
        # Add download link if requested
        if include_download_link and self.input_files:
            zip_path = self.create_zip_with_marked_files()
            if zip_path:
                html += f"""
                <div class="section">
                    <a href="{zip_path}" class="download-button">Download Marked Files</a>
                </div>
                """
        
        # Close the HTML
        html += """
        </body>
        </html>
        """
        
        # Write the HTML to file
        with open(output_path, 'w') as f:
            f.write(html)
            
        return output_path
    
    def _group_consecutive_weeks(self, weeks_data: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
        """Group consecutive weeks for more concise reporting."""
        if not weeks_data:
            return []
            
        # Sort weeks by week number
        sorted_weeks = sorted(weeks_data, key=lambda w: str(w['week']))
        
        groups = []
        current_group = [sorted_weeks[0]]
        
        for i in range(1, len(sorted_weeks)):
            current_week = sorted_weeks[i]
            last_week = current_group[-1]
            
            # Try to extract week numbers for comparison
            try:
                current_num = int(str(current_week['week']).split(' ')[-1])
                last_num = int(str(last_week['week']).split(' ')[-1])
                
                if current_num == last_num + 1:
                    # Consecutive weeks
                    current_group.append(current_week)
                else:
                    # Non-consecutive, start a new group
                    groups.append(current_group)
                    current_group = [current_week]
            except:
                # If we can't parse week numbers, treat as non-consecutive
                groups.append(current_group)
                current_group = [current_week]
        
        # Add the last group
        if current_group:
            groups.append(current_group)
            
        return groups
    
    def _generate_recommendations(self, critical_weeks: List[Dict[str, Any]], 
                                 tight_weeks: List[Dict[str, Any]],
                                 incompatible_waffles: List[str],
                                 compatibility_issues: List[str] = None) -> List[str]:
        """Generate recommendations based on feasibility issues."""
        compatibility_issues = compatibility_issues or []
        recommendations = []
        
        # Supply deficit recommendations
        if critical_weeks:
            total_deficit = sum(w['demand'] - w['supply'] for w in critical_weeks)
            recommendations.append(f"Increase supply by at least {total_deficit:,} pans in critical weeks")
            
            # Find weeks with excess supply
            excess_weeks = []
            for week in self.checker.data['weeks']:
                week_demand = self.checker.data['total_demand'].get(week, 0)
                week_supply = self.checker.data['total_supply'].get(week, 0)
                
                if week_supply > week_demand * 1.5:  # More than 50% excess
                    excess = week_supply - (week_demand * 1.1)  # Calculate transferable excess
                    if excess > 0:
                        excess_weeks.append({
                            'week': week,
                            'excess': excess
                        })
            
            if excess_weeks:
                # Sort by excess (descending)
                excess_weeks.sort(key=lambda w: w['excess'], reverse=True)
                top_week = excess_weeks[0]
                recommendations.append(
                    f"Shift {top_week['excess']:,.0f}+ pans from Week {top_week['week']} to deficit weeks")
        
        # Waffle compatibility recommendations
        if incompatible_waffles:
            if len(incompatible_waffles) <= 3:
                waffle_list = ', '.join([str(w) for w in incompatible_waffles])
                recommendations.append(f"Add compatible pan types for waffles: {waffle_list}")
            else:
                recommendations.append(f"Add compatible pan types for {len(incompatible_waffles)} problematic waffle types")
        
        # Compatibility-constrained supply recommendations
        if compatibility_issues:
            # Extract affected waffles from issues
            affected_waffles = set()
            for issue in compatibility_issues:
                try:
                    waffle_name = issue.split("waffle '")[1].split("'")[0]
                    affected_waffles.add(waffle_name)
                except:
                    pass
            
            if affected_waffles:
                if len(affected_waffles) <= 3:
                    waffle_list = ', '.join(sorted([str(w) for w in affected_waffles]))
                    recommendations.append(f"Increase supply of compatible pans for waffles: {waffle_list}")
                else:
                    recommendations.append(f"Increase supply of compatible pans for {len(affected_waffles)} waffles with compatibility issues")
            
            recommendations.append("Consider diversifying pan compatibility for waffles with limited options")
        
        return recommendations
    
    def _locate_cells(self, sheet, point: Dict[str, Any]) -> List[Any]:
        """Locate cells in a worksheet based on the critical point information."""
        cells = []
        
        if point['cell_type'] == 'row':
            # Find row with the given week identifier
            week_identifier = point['identifier']
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), 1):
                if str(week_identifier) in str(row[0]):  # Assuming week is in the first column
                    # Mark all non-empty cells in this row
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cells.append(cell)
                    break
        
        elif point['cell_type'] == 'waffle':
            # Find row with the given waffle identifier
            waffle_identifier = point['identifier']
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), 1):
                if str(waffle_identifier) in str(row[0]):  # Assuming waffle is in the first column
                    # Mark all non-empty cells in this row
                    for col_idx in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cells.append(cell)
                    break
        
        return cells


# Testing and demonstration
if __name__ == "__main__":
    # Sample data for demonstration
    from feasibility_check import FeasibilityChecker
    
    # Create sample data
    data = {
        'weeks': ['Week 1', 'Week 2', 'Week 3', 'Week 4'],
        'waffle_types': ['Waffle A', 'Waffle B', 'Waffle C'],
        'pan_types': ['Pan X', 'Pan Y', 'Pan Z'],
        'total_demand': {'Week 1': 100, 'Week 2': 150, 'Week 3': 200, 'Week 4': 150},
        'total_supply': {'Week 1': 300, 'Week 2': 120, 'Week 3': 180, 'Week 4': 100},
        'allowed': {
            ('Waffle A', 'Pan X'): True,
            ('Waffle A', 'Pan Y'): True,
            ('Waffle B', 'Pan Y'): True,
            ('Waffle C', 'Pan Z'): False,
        },
        'demand': {},
        'supply': {},
        'cost': {},
        'wpp': {}
    }
    
    # Initialize the checker and run check
    checker = FeasibilityChecker(data)
    checker.check_feasibility()
    
    # Create the reporter
    input_files = {
        'PanSupply.xlsx': 'PanSupply.xlsx',
        'WafflePanCombinations.xlsx': 'WafflePanCombinations.xlsx'
    }
    reporter = FeasibilityReporter(checker, input_files)
    
    # Generate and print text report
    text_report = reporter.generate_text_report()
    print(text_report)
    
    # Generate HTML report
    html_path = reporter.generate_html_report()
    print(f"HTML report saved to: {html_path}")
    
    # Create marked files
    marked_files = reporter.generate_marked_excel_files()
    print(f"Marked files: {marked_files}") 